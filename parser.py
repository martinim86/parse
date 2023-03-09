from create_table import Table
import openpyxl
import datetime
from datetime import datetime, timedelta

class Parser:
    table = None

    def __init__(self):
       self.table = Table()

    def read_from_file(self):
        wookbook = openpyxl.load_workbook('./Приложение_к_заданию_бек_разработчика.xlsx')
        worksheet = wookbook.active

        for i in range(3, worksheet.max_row):
            sps = []
            for col in worksheet.iter_cols(1, worksheet.max_column - 2):
                sps.append(col[i].value)
            sps.append(datetime.today() - timedelta(days=i))
            self.insert_table(sps)


    def insert_table(self, row):
        self.table.insert_into_table(row)
    def select_from_table(self):
        self.table.select_total()



a = Parser()
a.read_from_file()
a.select_from_table()
